#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
改进版三语对齐工具：使用段落级别对齐和句子级别对齐相结合的方法，提高三语对齐的精度
"""

import os
import re
import argparse
import pandas as pd
import numpy as np
from tqdm import tqdm
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sentence_transformers import SentenceTransformer

def load_text_file(file_path):
    """加载文本文件并返回内容"""
    print(f"加载文本文件: {file_path}")
    with open(file_path, 'r', encoding='utf-8') as f:
        text = f.read()
    return text

def split_into_paragraphs(text):
    """将文本分割为段落"""
    print("将文本分割为段落...")
    # 按空行分割段落
    paragraphs = re.split(r'\n\s*\n', text)
    paragraphs = [p.strip() for p in paragraphs if p.strip()]
    print(f"共分割出 {len(paragraphs)} 个段落")
    return paragraphs

def split_into_sentences(text):
    """将文本分割为句子"""
    print("将文本分割为句子...")
    # 使用简单的规则分割句子
    sentences = re.split(r'([。！？!?])', text)
    
    # 将分割符与前面的文本重新组合
    temp_sentences = []
    for i in range(0, len(sentences) - 1, 2):
        if i + 1 < len(sentences):
            temp_sentences.append(sentences[i] + sentences[i + 1])
        else:
            temp_sentences.append(sentences[i])
    
    # 如果最后一个元素不是空的，添加它
    if len(sentences) % 2 == 1 and sentences[-1]:
        temp_sentences.append(sentences[-1])
    
    sentences = [s.strip() for s in temp_sentences if s.strip()]
    print(f"共分割出 {len(sentences)} 个句子")
    return sentences

def encode_sentences(sentences, model):
    """使用预训练模型编码句子"""
    print("编码句子...")
    encodings = model.encode(sentences, show_progress_bar=True)
    return encodings

def compute_similarity_matrix(encodings1, encodings2):
    """计算两组句子之间的相似度矩阵"""
    print("计算相似度矩阵...")
    # 归一化编码
    normalized_encodings1 = encodings1 / np.linalg.norm(encodings1, axis=1, keepdims=True)
    normalized_encodings2 = encodings2 / np.linalg.norm(encodings2, axis=1, keepdims=True)
    
    # 计算余弦相似度
    similarity_matrix = np.dot(normalized_encodings1, normalized_encodings2.T)
    
    return similarity_matrix

def align_texts_dp(similarity_matrix, gap_penalty=-0.2):
    """使用动态规划算法对齐两个文本"""
    print("使用动态规划算法对齐文本...")
    m, n = similarity_matrix.shape
    
    # 初始化得分矩阵和回溯矩阵
    score_matrix = np.zeros((m + 1, n + 1))
    backtrack = np.zeros((m + 1, n + 1), dtype=int)
    
    # 初始化第一行和第一列
    for i in range(1, m + 1):
        score_matrix[i, 0] = score_matrix[i-1, 0] + gap_penalty
        backtrack[i, 0] = 1  # 向上
    
    for j in range(1, n + 1):
        score_matrix[0, j] = score_matrix[0, j-1] + gap_penalty
        backtrack[0, j] = 2  # 向左
    
    # 填充得分矩阵
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            match = score_matrix[i-1, j-1] + similarity_matrix[i-1, j-1]
            delete = score_matrix[i-1, j] + gap_penalty
            insert = score_matrix[i, j-1] + gap_penalty
            
            max_score = max(match, delete, insert)
            score_matrix[i, j] = max_score
            
            if max_score == match:
                backtrack[i, j] = 0  # 对角线
            elif max_score == delete:
                backtrack[i, j] = 1  # 向上
            else:
                backtrack[i, j] = 2  # 向左
    
    # 回溯找到最佳对齐路径
    alignments = []
    i, j = m, n
    
    while i > 0 or j > 0:
        if i > 0 and j > 0 and backtrack[i, j] == 0:  # 对角线
            alignments.append((i-1, j-1, similarity_matrix[i-1, j-1]))
            i -= 1
            j -= 1
        elif i > 0 and backtrack[i, j] == 1:  # 向上
            alignments.append((i-1, None, 0))
            i -= 1
        else:  # 向左
            alignments.append((None, j-1, 0))
            j -= 1
    
    # 反转对齐结果，使其按照原始顺序排列
    alignments.reverse()
    
    return alignments

def parse_alignment_file(file_path):
    """解析对齐结果文件，提取句对和相似度"""
    print(f"解析对齐结果文件: {file_path}")
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 使用正则表达式提取每个对齐块
    pattern = r'相似度: ([\d\.-]+).*?\n中文: (.*?)\n日文: (.*?)(?:\n\n|\Z)'
    matches = re.findall(pattern, content, re.DOTALL)
    
    alignments = []
    for similarity, zh, ja in matches:
        alignments.append({
            'similarity': float(similarity),
            'zh': zh.strip(),
            'ja': ja.strip()
        })
    
    print(f"共解析出 {len(alignments)} 个对齐句对")
    return alignments

def align_paragraphs(zh1_paragraphs, zh2_paragraphs, model):
    """对齐两个中文版本的段落"""
    print("对齐段落...")
    
    # 编码段落
    zh1_encodings = encode_sentences(zh1_paragraphs, model)
    zh2_encodings = encode_sentences(zh2_paragraphs, model)
    
    # 计算相似度矩阵
    similarity_matrix = compute_similarity_matrix(zh1_encodings, zh2_encodings)
    
    # 使用动态规划算法对齐段落
    alignments = align_texts_dp(similarity_matrix, gap_penalty=-0.3)
    
    return alignments

def align_sentences_within_paragraphs(zh1_paragraph, zh2_paragraph, model):
    """在段落内对齐句子"""
    # 分割句子
    zh1_sentences = split_into_sentences(zh1_paragraph)
    zh2_sentences = split_into_sentences(zh2_paragraph)
    
    if not zh1_sentences or not zh2_sentences:
        return []
    
    # 编码句子
    zh1_encodings = encode_sentences(zh1_sentences, model)
    zh2_encodings = encode_sentences(zh2_sentences, model)
    
    # 计算相似度矩阵
    similarity_matrix = compute_similarity_matrix(zh1_encodings, zh2_encodings)
    
    # 使用动态规划算法对齐句子
    alignments = align_texts_dp(similarity_matrix, gap_penalty=-0.2)
    
    # 转换为句子对齐结果
    sentence_alignments = []
    for i, j, sim in alignments:
        if i is not None and j is not None:
            sentence_alignments.append({
                'zh1': zh1_sentences[i],
                'zh2': zh2_sentences[j],
                'similarity': sim
            })
    
    return sentence_alignments

def improved_trilingual_alignment(zh1_file, zh2_file, zh_ja_alignment_file):
    """改进的三语对齐方法"""
    print("执行改进的三语对齐...")
    
    # 加载文本文件
    zh1_text = load_text_file(zh1_file)
    zh2_text = load_text_file(zh2_file)
    
    # 分割段落
    zh1_paragraphs = split_into_paragraphs(zh1_text)
    zh2_paragraphs = split_into_paragraphs(zh2_text)
    
    # 加载预训练模型
    model = SentenceTransformer('LaBSE')
    
    # 段落级别对齐
    paragraph_alignments = align_paragraphs(zh1_paragraphs, zh2_paragraphs, model)
    
    # 解析中日对齐结果
    zh_ja_alignments = parse_alignment_file(zh_ja_alignment_file)
    
    # 创建中文1到日文的映射
    zh1_to_ja_map = {}
    for alignment in zh_ja_alignments:
        zh_text = alignment['zh']
        ja_text = alignment['ja']
        similarity = alignment['similarity']
        zh1_to_ja_map[zh_text] = {'ja': ja_text, 'similarity': similarity}
    
    # 在段落内对齐句子并合并结果
    trilingual_alignments = []
    
    for i, j, sim in paragraph_alignments:
        if i is not None and j is not None:
            # 对齐段落内的句子
            sentence_alignments = align_sentences_within_paragraphs(zh1_paragraphs[i], zh2_paragraphs[j], model)
            
            # 合并中日对齐结果
            for alignment in sentence_alignments:
                zh1_text = alignment['zh1']
                zh2_text = alignment['zh2']
                zh1_zh2_similarity = alignment['similarity']
                
                # 查找对应的日文句子
                ja_text = None
                zh1_ja_similarity = 0
                
                if zh1_text in zh1_to_ja_map:
                    ja_text = zh1_to_ja_map[zh1_text]['ja']
                    zh1_ja_similarity = zh1_to_ja_map[zh1_text]['similarity']
                
                trilingual_alignments.append({
                    'zh1': zh1_text,
                    'zh2': zh2_text,
                    'ja': ja_text,
                    'zh1_zh2_similarity': zh1_zh2_similarity,
                    'zh1_ja_similarity': zh1_ja_similarity
                })
    
    return trilingual_alignments

def create_excel(trilingual_alignments, output_file, low_threshold=0.4, medium_threshold=0.6):
    """创建Excel文件，使用不同颜色标识不同类别的句对"""
    print(f"创建Excel文件: {output_file}")
    
    # 创建DataFrame
    df = pd.DataFrame([
        {
            '序号': i + 1,
            '中文版本1': a['zh1'],
            '中文版本2': a['zh2'],
            '日文': a['ja'] if a['ja'] else '',
            '中文1-中文2相似度': a['zh1_zh2_similarity'],
            '中文1-日文相似度': a['zh1_ja_similarity'],
            '检查类别': get_category(a['zh1_zh2_similarity'], a['zh1_ja_similarity'], low_threshold, medium_threshold)
        } for i, a in enumerate(trilingual_alignments)
    ])
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "三语对齐结果"
    
    # 添加表头
    headers = ['序号', '中文版本1', '中文版本2', '日文', '中文1-中文2相似度', '中文1-日文相似度', '检查类别']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    ws.column_dimensions[get_column_letter(1)].width = 10  # 序号
    ws.column_dimensions[get_column_letter(2)].width = 50  # 中文版本1
    ws.column_dimensions[get_column_letter(3)].width = 50  # 中文版本2
    ws.column_dimensions[get_column_letter(4)].width = 50  # 日文
    ws.column_dimensions[get_column_letter(5)].width = 15  # 中文1-中文2相似度
    ws.column_dimensions[get_column_letter(6)].width = 15  # 中文1-日文相似度
    ws.column_dimensions[get_column_letter(7)].width = 20  # 检查类别
    
    # 添加数据并设置颜色
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 浅红色
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 浅黄色
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 浅绿色
    
    for i, row in enumerate(df.itertuples(), 2):
        ws.cell(row=i, column=1).value = row.序号
        ws.cell(row=i, column=2).value = row.中文版本1
        ws.cell(row=i, column=3).value = row.中文版本2
        ws.cell(row=i, column=4).value = row.日文
        ws.cell(row=i, column=5).value = df.iloc[i-2]['中文1-中文2相似度']
        ws.cell(row=i, column=6).value = df.iloc[i-2]['中文1-日文相似度']
        ws.cell(row=i, column=7).value = row.检查类别
        
        # 设置单元格对齐方式
        for col in range(1, 8):
            ws.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)
        
        # 根据类别设置颜色
        fill = None
        if row.检查类别 == "需要重点检查":
            fill = red_fill
        elif row.检查类别 == "建议检查":
            fill = yellow_fill
        else:
            fill = green_fill
        
        for col in range(1, 8):
            ws.cell(row=i, column=col).fill = fill
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"Excel文件已保存: {output_file}")

def get_category(zh1_zh2_similarity, zh1_ja_similarity, low_threshold=0.4, medium_threshold=0.6):
    """根据相似度确定检查类别"""
    # 如果日文相似度为0（表示没有对应的日文句子），则需要重点检查
    if zh1_ja_similarity == 0:
        return "需要重点检查"
    
    # 计算平均相似度
    avg_similarity = (zh1_zh2_similarity + zh1_ja_similarity) / 2
    
    if avg_similarity < low_threshold:
        return "需要重点检查"
    elif avg_similarity < medium_threshold:
        return "建议检查"
    else:
        return "可能不需要检查"

def main():
    parser = argparse.ArgumentParser(description='改进版三语对齐工具：使用段落级别对齐和句子级别对齐相结合的方法，提高三语对齐的精度')
    parser.add_argument('--zh1', required=True, help='第一个中文版本文件路径')
    parser.add_argument('--zh2', required=True, help='第二个中文版本文件路径')
    parser.add_argument('--zh-ja-alignment', required=True, help='已有的中日对齐结果文件路径')
    parser.add_argument('--output', required=True, help='输出Excel文件路径')
    parser.add_argument('--low-threshold', type=float, default=0.4, help='低相似度阈值 (默认: 0.4)')
    parser.add_argument('--medium-threshold', type=float, default=0.6, help='中等相似度阈值 (默认: 0.6)')
    
    args = parser.parse_args()
    
    # 执行改进的三语对齐
    trilingual_alignments = improved_trilingual_alignment(args.zh1, args.zh2, args.zh_ja_alignment)
    
    # 创建Excel文件
    create_excel(trilingual_alignments, args.output, args.low_threshold, args.medium_threshold)

if __name__ == "__main__":
    main()
