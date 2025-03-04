#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
将对齐结果转换为Excel格式，并根据相似度阈值标记需要检查的句对
"""

import re
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def parse_alignment_file(file_path):
    """解析对齐结果文件，提取句对和相似度"""
    print(f"解析对齐结果文件: {file_path}")
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 使用正则表达式提取每个对齐块
    pattern = r'相似度: ([\d\.-]+).*?\n中文: (.*?)\n日文: (.*?)(?:\n\n|\Z)'
    matches = re.findall(pattern, content, re.DOTALL)
    
    alignments = []
    for similarity, zh, ja in matches:
        alignments.append({
            'similarity': float(similarity),
            'zh': zh.strip(),
            'ja': ja.strip()
        })
    
    print(f"共解析出 {len(alignments)} 个对齐句对")
    return alignments

def filter_and_mark_alignments(aligned_pairs, low_threshold=0.4, medium_threshold=0.6):
    """根据相似度阈值过滤和标记对齐结果"""
    print("根据相似度阈值过滤和标记对齐结果...")
    
    filtered_results = []
    
    for i, pair in enumerate(aligned_pairs):
        similarity = pair['similarity']
        category = ""
        
        if similarity < low_threshold:
            category = "需要重点检查"
        elif similarity < medium_threshold:
            category = "建议检查"
        else:
            category = "可能不需要检查"
        
        filtered_results.append({
            '序号': i + 1,
            '中文': pair['zh'],
            '日文': pair['ja'],
            '相似度': similarity,
            '检查类别': category
        })
    
    return filtered_results

def create_excel(filtered_results, output_file):
    """创建Excel文件，使用不同颜色标识不同类别的句对"""
    print(f"创建Excel文件: {output_file}")
    
    # 创建DataFrame
    df = pd.DataFrame(filtered_results)
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "对齐结果"
    
    # 添加表头
    headers = ['序号', '中文', '日文', '相似度', '检查类别']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列宽
    ws.column_dimensions[get_column_letter(1)].width = 10  # 序号
    ws.column_dimensions[get_column_letter(2)].width = 50  # 中文
    ws.column_dimensions[get_column_letter(3)].width = 50  # 日文
    ws.column_dimensions[get_column_letter(4)].width = 15  # 相似度
    ws.column_dimensions[get_column_letter(5)].width = 20  # 检查类别
    
    # 添加数据并设置颜色
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 浅红色
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 浅黄色
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 浅绿色
    
    for i, row in enumerate(df.itertuples(), 2):
        ws.cell(row=i, column=1).value = row.序号
        ws.cell(row=i, column=2).value = row.中文
        ws.cell(row=i, column=3).value = row.日文
        ws.cell(row=i, column=4).value = row.相似度
        ws.cell(row=i, column=5).value = row.检查类别
        
        # 设置单元格对齐方式
        for col in range(1, 6):
            ws.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=True)
        
        # 根据类别设置颜色
        fill = None
        if row.检查类别 == "需要重点检查":
            fill = red_fill
        elif row.检查类别 == "建议检查":
            fill = yellow_fill
        else:
            fill = green_fill
        
        for col in range(1, 6):
            ws.cell(row=i, column=col).fill = fill
    
    # 保存Excel文件
    wb.save(output_file)
    print(f"Excel文件已保存: {output_file}")

def analyze_and_print_statistics(filtered_results):
    """分析并打印统计信息"""
    print("\n===== 统计信息 =====")
    
    total_pairs = len(filtered_results)
    
    # 计算相似度统计
    similarities = [pair['相似度'] for pair in filtered_results]
    avg_similarity = sum(similarities) / total_pairs if total_pairs > 0 else 0
    
    # 计算各类别的数量
    categories = {}
    for pair in filtered_results:
        category = pair['检查类别']
        if category not in categories:
            categories[category] = 0
        categories[category] += 1
    
    # 打印统计信息
    print(f"总句对数: {total_pairs}")
    print(f"平均相似度: {avg_similarity:.4f}")
    
    for category, count in categories.items():
        percentage = (count / total_pairs) * 100 if total_pairs > 0 else 0
        print(f"{category}: {count} ({percentage:.2f}%)")

def main():
    parser = argparse.ArgumentParser(description='将对齐结果转换为Excel格式，并根据相似度阈值标记需要检查的句对')
    parser.add_argument('--input', required=True, help='输入对齐结果文件路径')
    parser.add_argument('--output', required=True, help='输出Excel文件路径')
    parser.add_argument('--low-threshold', type=float, default=0.4, help='低相似度阈值 (默认: 0.4)')
    parser.add_argument('--medium-threshold', type=float, default=0.6, help='中等相似度阈值 (默认: 0.6)')
    
    args = parser.parse_args()
    
    # 解析对齐结果文件
    aligned_pairs = parse_alignment_file(args.input)
    
    # 根据相似度阈值过滤和标记对齐结果
    filtered_results = filter_and_mark_alignments(aligned_pairs, args.low_threshold, args.medium_threshold)
    
    # 分析并打印统计信息
    analyze_and_print_statistics(filtered_results)
    
    # 创建Excel文件
    create_excel(filtered_results, args.output)

if __name__ == "__main__":
    main()
